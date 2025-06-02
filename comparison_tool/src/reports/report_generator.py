"""Report generation module for creating comparison reports."""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Union
import logging
import zipfile
import os

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self, output_dir: str):
        """
        Initialize the report generator.
        
        Args:
            output_dir: Directory to save generated reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Style configurations
        self.pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)

    def _style_header_row(self, worksheet, row_num: int):
        """Apply styling to header row."""
        for cell in worksheet[row_num]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')

    def generate_regression_report(self, comparison_results: Dict[str, Any]) -> str:
        """
        Generate Excel regression report with multiple tabs.
        
        Args:
            comparison_results: Dictionary containing comparison data
            
        Returns:
            Path to generated report
        """
        wb = openpyxl.Workbook()
        
        # Create AggregationCheck sheet
        self._create_aggregation_sheet(wb, comparison_results['column_summary'])
        
        # Create CountCheck sheet
        self._create_count_sheet(wb, comparison_results['row_counts'])
        
        # Create DistinctCheck sheet
        self._create_distinct_check_sheet(wb, comparison_results['distinct_values'])
        
        # Save workbook
        report_path = self.output_dir / f"Regression_Report_{self.timestamp}.xlsx"
        wb.save(str(report_path))
        
        return str(report_path)

    def _create_aggregation_sheet(self, workbook: openpyxl.Workbook, 
                                column_summary: Dict[str, Dict[str, Any]]):
        """Create the AggregationCheck sheet."""
        sheet = workbook.active
        sheet.title = "AggregationCheck"
        
        # Write headers
        headers = ['Column', 'Source Sum', 'Target Sum', 'Difference', 'Result']
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Write data
        for col, stats in column_summary.items():
            if 'source_sum' in stats:  # Only process numeric columns
                source_sum = stats['source_sum']
                target_sum = stats['target_sum']
                difference = source_sum - target_sum
                result = 'PASS' if abs(difference) < 0.0001 else 'FAIL'
                
                row = [col, source_sum, target_sum, difference, result]
                sheet.append(row)
                
                # Apply conditional formatting
                last_row = sheet.max_row
                sheet.cell(last_row, 5).fill = self.pass_fill if result == 'PASS' else self.fail_fill

        self._adjust_column_widths(sheet)

    def _create_count_sheet(self, workbook: openpyxl.Workbook, 
                          count_data: Dict[str, int]):
        """Create the CountCheck sheet."""
        sheet = workbook.create_sheet(title="CountCheck")
        
        # Write headers
        headers = ['Source Name', 'Target Name', 'Source Count', 'Target Count', 
                  'Difference', 'Result']
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Write data
        source_count = count_data['source_count']
        target_count = count_data['target_count']
        difference = source_count - target_count
        result = 'PASS' if difference == 0 else 'FAIL'
        
        row = [count_data['source_name'], count_data['target_name'], 
               source_count, target_count, difference, result]
        sheet.append(row)
        
        # Apply conditional formatting
        sheet.cell(2, 6).fill = self.pass_fill if result == 'PASS' else self.fail_fill
        
        self._adjust_column_widths(sheet)

    def _create_distinct_check_sheet(self, workbook: openpyxl.Workbook, 
                                   distinct_values: Dict[str, Dict[str, Any]]):
        """Create the DistinctCheck sheet."""
        sheet = workbook.create_sheet(title="DistinctCheck")
        
        # Write headers
        headers = ['Column', 'Source Distinct Count', 'Target Distinct Count', 
                  'Values Match', 'Result']
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Write data
        for col, data in distinct_values.items():
            source_count = data['source_count']
            target_count = data['target_count']
            values_match = data['matching']
            result = 'PASS' if values_match and source_count == target_count else 'FAIL'
            
            row = [col, source_count, target_count, str(values_match), result]
            sheet.append(row)
            
            # Apply conditional formatting
            last_row = sheet.max_row
            sheet.cell(last_row, 5).fill = self.pass_fill if result == 'PASS' else self.fail_fill
        
        self._adjust_column_widths(sheet)

    def _adjust_column_widths(self, worksheet):
        """Adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def generate_difference_report(self, differences: pd.DataFrame) -> str:
        """
        Generate side-by-side difference report.
        
        Args:
            differences: DataFrame containing differences
            
        Returns:
            Path to generated report
        """
        if differences.empty:
            # Create report indicating no differences
            report_path = self.output_dir / f"DifferenceReport_{self.timestamp}.txt"
            with open(report_path, 'w') as f:
                f.write("There are No Differences found.")
            return str(report_path)
        
        # Create Excel report with differences
        report_path = self.output_dir / f"DifferenceReport_{self.timestamp}.xlsx"
        
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            differences.to_excel(writer, index=False, sheet_name='Differences')
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets['Differences']
            
            self._style_header_row(worksheet, 1)
            self._adjust_column_widths(worksheet)
        
        return str(report_path)

    def create_report_archive(self, report_paths: Dict[str, str]) -> str:
        """
        Create ZIP archive containing all reports.
        
        Args:
            report_paths: Dictionary mapping report types to their file paths
            
        Returns:
            Path to ZIP archive
        """
        archive_path = self.output_dir / f"ComparisonReports_{self.timestamp}.zip"
        
        with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for report_type, path in report_paths.items():
                if os.path.exists(path):
                    # Add file to archive with a descriptive name
                    arcname = f"{report_type}_{os.path.basename(path)}"
                    zipf.write(path, arcname)
        
        return str(archive_path)

    def cleanup_reports(self, report_paths: List[str]):
        """
        Clean up individual report files after creating archive.
        
        Args:
            report_paths: List of paths to report files
        """
        for path in report_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as e:
                logger.warning(f"Failed to remove report file {path}: {str(e)}")
